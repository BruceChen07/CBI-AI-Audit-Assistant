"""
Author: Bruce Chen <bruce.chen@effem.com>
Date: 2025-08-29

Copyright (c) 2025 Mars Corporation

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import pandas as pd
import io
import logging
import tempfile
from datetime import datetime
from fastapi import UploadFile, HTTPException
from typing import Dict, Any, List
from models import ExcelData, PDFUploadResponse
from rag_service import process_pdf
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles.colors import Color

logger = logging.getLogger(__name__)

class FileService:
    @staticmethod
    async def process_excel_file(file: UploadFile) -> ExcelData:
        """Process uploaded Excel file"""
        logger.info(f"Received file upload request: {file.filename}")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            logger.warning(f"Unsupported file format: {file.filename}")
            raise HTTPException(status_code=400, detail="Only accepts Excel file format(.xlsx, .xls)")
        
        try:
            logger.debug(f"Start reading the file: {file.filename}")
            contents = await file.read()
            logger.debug(f"File {file.filename} content read completed, size: {len(contents)} bytes")
            
            logger.debug(f"Start reading Excel file: {file.filename}")
            df = pd.read_excel(io.BytesIO(contents), skiprows=5)  # Parse from line 6
            logger.info(f"Excel file {file.filename} read completed, rows: {len(df)}, columns: {len(df.columns)}")
            
            required_columns = ["Chapter", "Element", "Criteria", "Hint", 
                              "Classification", "Comments", "Previous Comments", "AET"]
            
            logger.debug(f"Checking the required columns: {', '.join(required_columns)}")
            df_columns = [col.strip() for col in df.columns]
            logger.debug(f"the columns in the file: {', '.join(df_columns)}")
            
            missing_columns = [col for col in required_columns 
                             if not any(col.lower() == c.lower() for c in df_columns)]
            
            if missing_columns:
                logger.warning(f"File {file.filename} is missing required columns: {', '.join(missing_columns)}")
                raise HTTPException(
                    status_code=400, 
                    detail=f"Excel file {file.filename} is missing required columns: {', '.join(missing_columns)}"
                )
            
            logger.debug("Start converting DataFrame to dictionary list")
            data = df.to_dict(orient='records')
            logger.info(f"Data processing completed, {len(data)} records returned")
            
            return ExcelData(data=data)
        
        except Exception as e:
            logger.error(f"Failed to process file: {str(e)}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
    
    @staticmethod
    def generate_excel_from_cache(data: List[Dict[str, Any]], filename: str = None, statistics: Dict = None) -> str:
        """Generate Excel file from cached data (optimized version)"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f'audit_results_{timestamp}.xlsx'
            
            logger.info(f"Generating Excel from cached data: {len(data)} records")
            
            # Record field statistics
            field_stats = {}
            for row in data:
                for field in ['Evidence Collected by AI', 'Reference', 'AET Evidence Collected by AI', 'AET Reference']:
                    if field in row:
                        value = str(row[field]).strip()
                        if value and value != 'No data available':
                            field_stats[field] = field_stats.get(field, 0) + 1
            
            logger.info(f"Field statistics: {field_stats}")
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Ensure all required fields exist
            expected_fields = ['Evidence Collected by AI', 'Reference', 'AET Evidence Collected by AI', 'AET Reference']
            for field in expected_fields:
                if field not in df.columns:
                    df[field] = 'No data available'
                    logger.warning(f"Added missing field: {field}")
            
            # Create a temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file_path = temp_file.name
            temp_file.close()
            
            # Write Excel file
            with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Audit Results', index=False)
                
                # Get worksheet for formatting
                worksheet = writer.sheets['Audit Results']
                
                # Set header row style
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column width
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # === Keyword highlighting: apply user-defined keywords/colors from admin config ===
                # Load keyword configs from admin config storage
                try:
                    import auth as _auth
                    keyword_items = _auth.get_keyword_configs()  # [{'keyword': 'Findings', 'color': '#FF0000'}, ...]
                except Exception:
                    keyword_items = []

                # 富文本能力检查（不支持则报错）
                try:
                    from openpyxl.cell.rich_text import TextBlock, CellRichText
                    from openpyxl.cell.text import InlineFont
                except Exception as e:
                    raise RuntimeError("openpyxl>=3.1 且需支持 openpyxl.cell.rich_text 才能进行片段高亮") from e

                import re

                def _sanitize_hex(c: str) -> str:
                    s = str(c or "").strip().upper()
                    if s.startswith("#"):
                        s = s[1:]
                    return s if len(s) == 6 else "FF0000"

                # 构建忽略大小写的 keyword->color 映射与联合正则（长词优先）
                keyword_map: dict[str, str] = {}
                for it in keyword_items or []:
                    kw = str((it or {}).get("keyword", "")).strip()
                    color_hex = _sanitize_hex((it or {}).get("color", "FF0000"))
                    if kw:
                        keyword_map[kw.lower()] = color_hex

                pattern = None
                if keyword_map:
                    keys_sorted = sorted(keyword_map.keys(), key=len, reverse=True)
                    pattern = re.compile("|".join(map(re.escape, keys_sorted)), re.IGNORECASE)

                def _to_inline_font(cell_font) -> InlineFont:
                    # 修复：不使用未定义变量 f，不向构造函数传递不支持的 name 参数
                    base = cell_font
                    inline = InlineFont()
                    inline.sz = getattr(base, "sz", None) or getattr(base, "size", None) or 11
                    inline.b = bool(getattr(base, "bold", False))
                    inline.i = bool(getattr(base, "italic", False))
                    u_raw = getattr(base, "underline", None)
                    if isinstance(u_raw, bool):
                        inline.u = "single" if u_raw else "none"
                    elif not u_raw:
                        inline.u = "none"
                    else:
                        inline.u = str(u_raw).lower()
                    inline.strike = bool(getattr(base, "strike", False))
                    return inline

                def _clone_inline_font(f: InlineFont) -> InlineFont:
                    # 保留基础样式；不复制颜色到普通片段
                    g = InlineFont()
                    g.sz = getattr(f, "sz", None)
                    g.b = getattr(f, "b", None)
                    g.i = getattr(f, "i", None)
                    g.u = getattr(f, "u", None)
                    g.strike = getattr(f, "strike", None)
                    return g

                # 应用关键字高亮到所有数据行
                if pattern:
                    for row_idx in range(2, len(df) + 2):  # 从第2行开始（跳过表头）
                        for col_idx, col_name in enumerate(df.columns, 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell_value = str(cell.value or "")
                            
                            if cell_value and pattern.search(cell_value):
                                # 找到匹配的关键字，应用富文本高亮
                                blocks = []
                                last_end = 0
                                
                                for match in pattern.finditer(cell_value):
                                    # 添加匹配前的普通文本
                                    if match.start() > last_end:
                                        normal_text = cell_value[last_end:match.start()]
                                        blocks.append(TextBlock(_to_inline_font(cell.font), normal_text))
                                    
                                    # 添加高亮的关键字
                                    keyword = match.group().lower()
                                    color_hex = keyword_map.get(keyword, "FF0000")
                                    highlight_font = _clone_inline_font(_to_inline_font(cell.font))
                                    highlight_font.color = Color(rgb=color_hex)
                                    blocks.append(TextBlock(highlight_font, match.group()))
                                    
                                    last_end = match.end()
                                
                                # 添加剩余的普通文本
                                if last_end < len(cell_value):
                                    remaining_text = cell_value[last_end:]
                                    blocks.append(TextBlock(_to_inline_font(cell.font), remaining_text))
                                
                                # 应用富文本到单元格
                                if blocks:
                                    cell.value = CellRichText(blocks)

            logger.info(f"Excel file generated successfully: {temp_file_path}")
            return temp_file_path
            
        except Exception as e:
            logger.error(f"Error generating Excel from cache: {str(e)}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")
    
    @staticmethod
    async def process_pdf_file(file: UploadFile) -> PDFUploadResponse:
        """Process uploaded PDF file"""
        logger.info(f"Received PDF file upload request: {file.filename}")
        
        if not file.filename.lower().endswith('.pdf'):
            logger.warning(f"Format is not supported: {file.filename}")
            raise HTTPException(status_code=400, detail="Only support Pdf format(.pdf)")
        
        try:
            logger.debug(f"Start reading Pdf file: {file.filename}")
            pdf_bytes = await file.read()
            file_size_mb = len(pdf_bytes) / (1024 * 1024)
            logger.debug(f"Pdf file content read completed: {file.filename}, size: {file_size_mb:.2f} MB")
            
            logger.info(f"Start processing Pdf file: {file.filename}")
            result = process_pdf(pdf_bytes, file.filename)
            
            if result["success"]:
                logger.info(f"Pdf file processing completed successfully")
                return PDFUploadResponse(
                    success=True,
                    message="Pdf file upload and processing completed successfully",
                    documents_processed=result["documents_processed"]
                )
            else:
                logger.error(f"Pdf file processing failed: {result['error']}")
                return PDFUploadResponse(
                    success=False,
                    message="Pdf file processing failed",
                    error=result["error"]
                )
        
        except Exception as e:
            logger.error(f"Pdf file upload and processing failed: {str(e)}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Pdf file upload and processing failed: {str(e)}")